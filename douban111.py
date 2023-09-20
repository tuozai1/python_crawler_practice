'''获取豆瓣top250电影'''
import openpyxl
import time
import requests
from comp import *
from bs4 import BeautifulSoup

headers = {                 # 伪装为浏览器发送请求
     "User-agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36 Edg/116.0.1938.69"
}

class Douban:
    def __init__(self):
        self.headers = {                # 伪装为浏览器发送请求
            "User-agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/116.0.0.0 Safari/537.36 Edg/116.0.1938.69"
        }
        self.datelist = []          # 保存所有电影的信息

    def get_information(self):      # 获取信息
        for start_num in range(0, 250, 25):  # 网页为25个电影一页，步数25循环
            URL = f"https://movie.douban.com/top250?start={start_num}"  # 记录网址
            response = requests.get(URL, headers=headers)  # 获取每页信息

            if response.status_code == 200:
                html = response.text
                soup = BeautifulSoup(html, "html.parser")
                items = soup.findAll("div", class_="item")
                self.output_result(items)
            elif response.status_code == 403:
                print("被服务器禁止访问")
            else:
                print("未知异常")

    def output_result(self, items):     # 逐个保存
        for item in items:
            date = []  # 保存一部电影的所有信息
            item = str(item)
            self.date_append(date, item)

    def date_append(self, date, item):      # 保存电影信息
        link = re.findall(findLink, item)[0]    # 电影链接
        date.append(link)

        imgSrc = re.findall(findImgSrc, item)[0]    # 电影图片
        date.append(imgSrc)

        title = re.findall(findTitle, item)[0]      # 片名
        date.append(title)

        rating = re.findall(findRating, item)[0]    # 评分
        date.append(rating)

        JudgeNum = re.findall(findJudge, item)[0]   # 评价人数
        date.append(JudgeNum)

        inq = re.findall(findInq, item)             # 概况
        if len(inq) != 0:
            inq = inq[0].replace("。 ", "")
            date.append(inq)
        else:
            date.append(" ")                # 留空

        bd = re.findall(findBd, item)[0]        # 电影相关内容
        bd = re.sub('<br(\s+)?/>(\s+)?', " ", bd)
        date.append(bd.strip())             # 去掉空格

        self.datelist.append(date)

    def save_data(self):                    # 保存excel
        print("Saving data to Excel...")

        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = '豆瓣电影Top250'

        col = ("", "电影详情链接", "图片链接", "片名", "评分", "评价数", "概况", "相关信息")
        for i in range(1, 8):
            sheet.cell(1, i + 1).value = col[i]

        for i in range(0, 250):
            print(f'Processing {i + 1}th movie...')
            date = self.datelist[i]
            sheet.cell(i + 2, 1).value = i + 1
            for j in range(0, 7):
                sheet.cell(i + 2, j + 2).value = date[j]

        wb.save('豆瓣top250.xlsx')
        print("Data saved successfully.")

if __name__ == "__main__":
    time.sleep(1)
    execute = Douban()
    execute.get_information()
    execute.save_data()